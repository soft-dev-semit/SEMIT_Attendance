from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Color
from .models import Lesson_visit, Student
import os


def create_excel_template(discipline, groups):
    file_name = 'excel_template.xlsx'
    # Проверяем существует ли файл
    if os.path.exists(file_name):
        os.remove(file_name)
        print(f"Старый файл {file_name} удален.")

    # Создаем новую книгу Excel
    wb = Workbook()

    # Добавляем новый лист
    ws = wb.active

    ws.freeze_panes = 'D4'

    bold_font = Font(bold=True)
    font = Font(size=14)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    center_alignment = Alignment(horizontal='center')

    # fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws.append([''] * 3)
    ws.append([''] * 3)
    ws.append(['#', 'Група', 'Прізвище І.П.'])
    ws['C2'] = 'Номер лекції'
    ws['G1'] = 10
    ws['B1'] = discipline.abbrev

    # Применяем жирный шрифт и выравнивание по центру к заголовкам
    for cell in ws[3]:
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Создаем словарь для хранения номера строки для каждого студента
    student_numbers = {}

    # Записываем студентов в Excel-файл
    for group in groups:
        students = Student.objects.filter(group__name=group).order_by('group__name', 'id')
        for student in students:
            if student.id not in student_numbers:
                student_numbers[student.id] = len(student_numbers) + 4
            row_data = [len(student_numbers), student.group.name, student.full_name]
            for idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=len(student_numbers) + 3, column=idx, value=cell_value)
                cell.border = thin_border
                if idx != 3:
                    cell.alignment = center_alignment


    # Получаем все уникальные даты лекций по данной дисциплине и группам
    all_dates = Lesson_visit.objects.filter(discipline=discipline, group__name__in=groups).values_list('date',
                                                                                                       flat=True).distinct()
    column = 4
    for date in all_dates:
        formatted_date = date.strftime('%d/%m')
        ws.cell(row=3, column=column, value=formatted_date)
        ws.cell(row=3, column=column).alignment = center_alignment
        ws.cell(row=3, column=column).border = thin_border
        column += 1

    # Записываем "Всього" в следующей ячейке после дат
    ws.cell(row=3, column=column, value="Всього")
    ws.cell(row=3, column=column).alignment = center_alignment
    ws.cell(row=3, column=column).border = thin_border

    # Применяем границы к заголовкам и столбцам с датами и "Всього"
    for col in ws.iter_cols(min_col=1, max_col=column, min_row=3, max_row=3):
        for cell in col:
            cell.border = thin_border

    for idx, date in enumerate(all_dates, start=4):
        ws.cell(row=2, column=idx, value=idx - 3)

    # Посещаемость студентов
    for student in Student.objects.filter(group__name__in=groups):
        # Перебираем все даты лекций
        for column, date in enumerate(all_dates, start=4):
            # Проверяем, был ли студент на лекции в этот день
            visit_exists = Lesson_visit.objects.filter(email=student, discipline=discipline, date=date).exists()

            # Записываем результат посещения в ячейку
            ws.cell(row=student_numbers[student.id], column=column, value=int(visit_exists))
            ws.cell(row=student_numbers[student.id], column=column).alignment = center_alignment
            ws.cell(row=student_numbers[student.id], column=column).border = thin_border

        # Добавляем формулу для вычисления суммы посещений студента и округления вверх
        sum_formula = f'=ROUNDUP(SUM(D{student_numbers[student.id]}:F{student_numbers[student.id]})/COUNT(D{student_numbers[student.id]}:F{student_numbers[student.id]})*$G$1, 0)'
        ws.cell(row=student_numbers[student.id], column=column + 1, value=sum_formula)
        ws.cell(row=student_numbers[student.id], column=column + 1).border = thin_border

    wb.save(file_name)
    print(f"Шаблон Excel создан в файле: {file_name}")

    return os.path.abspath(file_name)
